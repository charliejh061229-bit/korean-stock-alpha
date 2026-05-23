import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

UNIVERSE = {
    "한미반도체":   "042700.KS",
    "ISC":          "095340.KQ",
    "리노공업":     "058470.KQ",
    "원익IPS":      "240810.KQ",
    "피에스케이":   "319660.KQ",
    "하나머티리얼즈": "166090.KQ",
    "HPSP":         "403870.KQ",
    "동진쎄미켐":   "005290.KQ",
    "솔브레인":     "357780.KQ",
    "이엔에프테크놀로지": "102710.KQ",
    "SK머티리얼즈": "036490.KQ",
    "티씨케이":     "064760.KQ",
    "원익머트리얼즈": "104830.KQ",
    "GST":          "083450.KQ",
    "서진시스템":   "178320.KQ",
    "코미코":       "183300.KQ",
    "주성엔지니어링": "036930.KQ",
    "유진테크":     "084370.KQ",
    "테스":         "095610.KQ",
    "케이씨텍":     "281404.KQ",
}

WEIGHTS = {"momentum": 0.40, "quality": 0.35, "lowvol": 0.25}

def fetch(name, ticker):
    try:
        s    = yf.Ticker(ticker)
        info = s.info
        hist = s.history(period="1y")
        if hist.empty or len(hist) < 21:
            return None

        close = hist["Close"]
        daily_ret = close.pct_change().dropna()

        p_now  = close.iloc[-1]
        p_1m   = close.iloc[-21]  if len(close) >= 21  else close.iloc[0]
        p_12m  = close.iloc[0]
        
        ret_12m_ex1m = (p_1m / p_12m) - 1          
        volatility   = daily_ret.std() * np.sqrt(252)
        risk_adj_mom = ret_12m_ex1m / volatility if volatility > 0 else 0

        
        roe_curr = info.get("returnOnEquity")
        eps_trail = info.get("trailingEps", 0) or 0
        eps_fwd   = info.get("forwardEps",  0) or 0
        earnings_trend = (eps_fwd - eps_trail) / abs(eps_trail) if eps_trail != 0 else 0

        return {
            "종목명":        name,
            "현재가":        round(p_now),
            "52주 수익률":   round((p_now / p_12m - 1) * 100, 2),
            "1M 수익률":     round((p_now / p_1m  - 1) * 100, 2),
            "연간 변동성":   round(volatility * 100, 2),
            "위험조정모멘텀": round(risk_adj_mom, 4),
            "ROE":           round(roe_curr * 100, 2) if roe_curr else None,
            "영업이익률":    round(info.get("operatingMargins", 0) * 100, 2) if info.get("operatingMargins") else None,
            "이익추세":      round(earnings_trend * 100, 2),
        }
    except:
        return None


def collect():
    print(f"총 {len(UNIVERSE)}개 종목 데이터 수집 중...\n")
    rows = []
    for i, (name, ticker) in enumerate(UNIVERSE.items(), 1):
        print(f"  [{i:02d}/{len(UNIVERSE)}] {name}", end=" ", flush=True)
        d = fetch(name, ticker)
        if d:
            rows.append(d)
            print("✓")
        else:
            print("✗")
    return pd.DataFrame(rows)

def rank_pct(series, ascending=True):
    return series.rank(ascending=ascending, pct=True).mul(100).round(1)

def score(df):
    df = df.copy()

    df["모멘텀스코어"] = rank_pct(df["위험조정모멘텀"], ascending=False)

    roe_s    = rank_pct(df["ROE"].fillna(df["ROE"].median()),                 ascending=False)
    opm_s    = rank_pct(df["영업이익률"].fillna(df["영업이익률"].median()),   ascending=False)
    trend_s  = rank_pct(df["이익추세"],                                       ascending=False)
    df["퀄리티스코어"] = ((roe_s + opm_s + trend_s) / 3).round(1)
    df["저변동성스코어"] = rank_pct(df["연간 변동성"], ascending=True)

    df["알파스코어"] = (
        df["모멘텀스코어"]  * WEIGHTS["momentum"] +
        df["퀄리티스코어"]  * WEIGHTS["quality"]  +
        df["저변동성스코어"] * WEIGHTS["lowvol"]
    ).round(1)

    df["시그널"] = df["알파스코어"].apply(
        lambda s: "▲ 매수" if s >= 65 else ("▼ 매도" if s <= 35 else "─ 중립")
    )
    return df.sort_values("알파스코어", ascending=False).reset_index(drop=True)

H_FILL  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
BUY_F   = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
SELL_F  = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
THIN    = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"),  bottom=Side(style="thin"))

def hcell(ws, r, c, val, w=None):
    cl = ws.cell(r, c, val)
    cl.font      = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    cl.fill      = H_FILL
    cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cl.border    = THIN
    if w:
        ws.column_dimensions[get_column_letter(c)].width = w

def dcell(ws, r, c, val, fmt=None, fill=None, bold=False):
    cl = ws.cell(r, c, val)
    cl.font      = Font(name="Arial", size=9, bold=bold)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    cl.border    = THIN
    if fmt:  cl.number_format = fmt
    if fill: cl.fill = fill

def export(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "알파 스코어"
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:N1")
    ws.merge_cells("A2:N2")
    ws["A1"].value     = "KOSPI/KOSDAQ 멀티팩터 알파 스코어링"
    ws["A1"].font      = Font(bold=True, size=14, name="Arial", color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].value     = (f"생성일: {datetime.today().strftime('%Y-%m-%d %H:%M')}  |  "
                          f"모멘텀 40% (위험조정) + 퀄리티 35% (ROE·OPM·이익추세) + 저변동성 25%  |  총 {len(df)}개 종목")
    ws["A2"].font      = Font(size=9, name="Arial", color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16

    headers = [
        ("순위",7), ("종목명",14), ("현재가",11),
        ("52주\n수익률",9), ("1M\n수익률",9), ("연간\n변동성",9),
        ("ROE",8), ("영업\n이익률",9), ("이익\n추세",9),
        ("모멘텀\n스코어",9), ("퀄리티\n스코어",9), ("저변동성\n스코어",10),
        ("알파\n스코어",10), ("시그널",9),
    ]
    for ci, (h, w) in enumerate(headers, 1):
        hcell(ws, 3, ci, h, w)
    ws.row_dimensions[3].height = 28

    cols = ["종목명","현재가","52주 수익률","1M 수익률","연간 변동성",
            "ROE","영업이익률","이익추세",
            "모멘텀스코어","퀄리티스코어","저변동성스코어","알파스코어","시그널"]
    fmts = [None,"#,##0","0.00%","0.00%","0.0%",
            "0.0%","0.0%","0.0%",
            "0.0","0.0","0.0","0.0",None]

    for ri, row in df.iterrows():
        er  = ri + 4
        sig = row["시그널"]
        rf  = BUY_F if "매수" in sig else (SELL_F if "매도" in sig else None)

        dcell(ws, er, 1, ri + 1)
        for ci, (col, fmt) in enumerate(zip(cols, fmts), 2):
            val = row[col]
            if col in ("52주 수익률","1M 수익률","연간 변동성","ROE","영업이익률","이익추세"):
                val = val / 100 if val is not None else None
            fill = rf if ci >= 10 else None
            dcell(ws, er, ci, val, fmt=fmt, fill=fill, bold=(col=="알파스코어"))
        ws.row_dimensions[er].height = 18

    lr = len(df) + 5
    ws.cell(lr, 1, " 팩터 정의").font = Font(bold=True, name="Arial", size=9)
    notes = [
        ("모멘텀 (40%)", "12M-1M 수익률 ÷ 연간 변동성  →  과열 종목 제거 + 위험 조정"),
        ("퀄리티 (35%)", "ROE + 영업이익률 + EPS 이익추세 (YoY 개선) 평균"),
        ("저변동성 (25%)", "52주 일간 수익률 표준편차 역수  →  변동성 낮을수록 고점수"),
    ]
    for i, (f, d) in enumerate(notes):
        r = lr + 1 + i
        ws.cell(r, 1, f).font = Font(bold=True, name="Arial", size=9)
        ws.cell(r, 2, d).font = Font(name="Arial", size=9, color="595959")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

    fname = f"alpha_score_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = os.path.join(os.getcwd(), fname)
    wb.save(fpath)
    return fpath

if __name__ == "__main__":
    df = collect()
    if df.empty:
        print("\n데이터 수집 실패.")
    else:
        df = score(df)
        print(f"\n── 상위 10개 ─────────────────────────────────────────────")
        print(df[["종목명","모멘텀스코어","퀄리티스코어","저변동성스코어","알파스코어","시그널"]].head(10).to_string(index=False))
        path = export(df)
        print(f"\n✅ 저장 완료: {path}")
